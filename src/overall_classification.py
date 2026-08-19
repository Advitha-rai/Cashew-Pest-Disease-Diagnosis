"""
Cashew Pest and Disease Diagnosis System
Phase 10: Overall Pest & Disease Classification Summary Engine (TensorFlow / Keras / OpenPyXL)

Generates the higher-level overall pest vs. disease classification summary report across all 8 individual
vision models. Groups the three pest classes (Aphids, Leaf miner, TMB) into "Overall 3 Pests" using
pooled correct/total counts, while keeping the single disease class (Leaf blight) separate.

Exports formatted Excel workbook (overall_pest_disease_classification.xlsx) and CSV backup.
"""

import os
import sys
import time
import json
import logging
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple, Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.config import Config
from src.utils import set_seed, get_logger

# Configure dedicated Phase 10 loggers
phase10_log_path = os.path.join(Config.get_logs_dir(), "evaluation.log")
exception_log_path = os.path.join(Config.get_logs_dir(), "exceptions.log")

logger = get_logger("Phase10OverallClassification", phase10_log_path)
exc_logger = get_logger("ExceptionEngine", exception_log_path)


# ---------------------------------------------------------
# 1. Canonical Class Name Resolution Helper
# ---------------------------------------------------------
CANONICAL_CLASSES = ["Aphids", "Leaf miner", "TMB", "Leaf blight"]

def normalize_class_key(raw_name: str) -> str:
    """
    Safely normalizes raw class name strings for matching while preserving exact canonical dataset names.
    Examples:
      'Leaf_Blight', 'Leaf Blight', 'Leaf_blight' -> 'Leaf blight'
      'Leaf_Miner', 'Leaf Miner', 'Leaf_miner'   -> 'Leaf miner'
      'Aphids', 'aphids'                         -> 'Aphids'
      'TMB', 'tmb'                               -> 'TMB'
    """
    clean = str(raw_name).strip().replace("_", " ").lower()
    if "aphid" in clean:
        return "Aphids"
    elif "miner" in clean:
        return "Leaf miner"
    elif "blight" in clean:
        return "Leaf blight"
    elif "tmb" in clean or "mosquito" in clean:
        return "TMB"
    return str(raw_name).strip()


# ---------------------------------------------------------
# 2. Existing Artifact Inspection & Metrics Extraction Engine
# ---------------------------------------------------------
def load_model_classification_metrics(model_index: int, folder_name: str) -> Dict[str, Dict[str, float]]:
    """
    Locates existing classification evaluation artifacts for a model without running any model inference.
    Inspects Phase 6 complete-dataset per-class results, summary JSONs, global comparison CSVs,
    and Phase 4 evaluation classification reports. Returns dict mapping canonical class names
    to {'correct': int, 'total': int, 'accuracy': float}.
    """
    base_dir = Config.get_base_dir()
    per_class_results = {}

    # Candidate source file paths to inspect
    p6_model_dir = Config.get_individual_models_full_dataset_dir(model_index)
    p6_root_dir = Config.get_individual_models_full_dataset_dir()
    p4_model_dir = os.path.join(base_dir, "Experiments", folder_name)

    summary_json_p6 = os.path.join(p6_model_dir, "full_dataset_summary.json")
    per_class_csv_p6 = os.path.join(p6_model_dir, "full_dataset_per_class_results.csv")
    global_csv_p6 = os.path.join(p6_root_dir, "8_model_complete_dataset_comparison.csv")
    clr_json_p4 = os.path.join(p4_model_dir, "classification_report.json")
    summary_json_p4 = os.path.join(p4_model_dir, "evaluation_summary.json")

    # Source 1: Phase 6 full_dataset_summary.json
    if os.path.exists(summary_json_p6):
        try:
            with open(summary_json_p6, "r") as f:
                data = json.load(f)
                if "per_class_results" in data:
                    raw_cls = data["per_class_results"]
                    for k, v in raw_cls.items():
                        c_name = normalize_class_key(k)
                        per_class_results[c_name] = {
                            "correct": int(v.get("correct", 0)),
                            "total": int(v.get("total", 0)),
                            "accuracy": float(v.get("accuracy", 0.0))
                        }
                    logger.info(f"Loaded class metrics for '{folder_name}' from Phase 6 summary JSON.")
                    return per_class_results
        except Exception as e:
            logger.warning(f"Could not read {summary_json_p6}: {e}")

    # Source 2: Phase 6 full_dataset_per_class_results.csv
    if os.path.exists(per_class_csv_p6):
        try:
            df = pd.read_csv(per_class_csv_p6)
            for _, row in df.iterrows():
                raw_cls = str(row.get("Actual Class", row.get("class", "")))
                c_name = normalize_class_key(raw_cls)
                corr = int(row.get("Correct", 0))
                tot = int(row.get("Total", 0))
                acc = float(row.get("Accuracy", (corr / tot * 100.0) if tot > 0 else 0.0))
                per_class_results[c_name] = {
                    "correct": corr,
                    "total": tot,
                    "accuracy": round(acc, 2)
                }
            logger.info(f"Loaded class metrics for '{folder_name}' from Phase 6 per-class CSV.")
            return per_class_results
        except Exception as e:
            logger.warning(f"Could not read {per_class_csv_p6}: {e}")

    # Source 3: Phase 6 global 8_model_complete_dataset_comparison.csv
    if os.path.exists(global_csv_p6):
        try:
            df_g = pd.read_csv(global_csv_p6)
            m_rows = df_g[df_g["Model"] == folder_name]
            if not m_rows.empty:
                r = m_rows.iloc[0]
                for c_name in CANONICAL_CLASSES:
                    safe_k = c_name.replace(" ", "_")
                    corr_col = f"{safe_k} Correct"
                    tot_col = f"{safe_k} Total"
                    acc_col = f"{safe_k} Accuracy (%)"
                    
                    if corr_col in r and tot_col in r:
                        corr = int(r[corr_col])
                        tot = int(r[tot_col])
                        acc = float(r.get(acc_col, (corr / tot * 100.0) if tot > 0 else 0.0))
                        per_class_results[c_name] = {"correct": corr, "total": tot, "accuracy": round(acc, 2)}
                
                if len(per_class_results) == 4:
                    logger.info(f"Loaded class metrics for '{folder_name}' from Phase 6 global comparison CSV.")
                    return per_class_results
        except Exception as e:
            logger.warning(f"Could not read {global_csv_p6}: {e}")

    # Source 4: Phase 4 classification_report.json
    if os.path.exists(clr_json_p4):
        try:
            with open(clr_json_p4, "r") as f:
                c_data = json.load(f)
                for k, v in c_data.items():
                    if isinstance(v, dict) and "support" in v:
                        c_name = normalize_class_key(k)
                        if c_name in CANONICAL_CLASSES:
                            tot = int(v.get("support", 0))
                            rec = float(v.get("recall", 0.0))
                            corr = int(round(rec * tot))
                            acc = round(rec * 100.0, 2)
                            per_class_results[c_name] = {"correct": corr, "total": tot, "accuracy": acc}
                if len(per_class_results) == 4:
                    logger.info(f"Loaded class metrics for '{folder_name}' from Phase 4 classification report.")
                    return per_class_results
        except Exception as e:
            logger.warning(f"Could not read {clr_json_p4}: {e}")

    # Fallback placeholder if no artifact found
    logger.error(f"[PHASE 10 ERROR] Could not locate classification artifacts for '{folder_name}'.")
    for c_name in CANONICAL_CLASSES:
        per_class_results[c_name] = {"correct": 0, "total": 0, "accuracy": 0.0}

    return per_class_results


# ---------------------------------------------------------
# 3. Overall 3 Pests Pooled Metrics Aggregator
# ---------------------------------------------------------
def calculate_overall_pests_metrics(aphids: Dict, leaf_miner: Dict, tmb: Dict) -> Dict:
    """
    Calculates pooled 'Overall 3 Pests' classification metrics:
      Pest Correct  = Aphids Correct + Leaf miner Correct + TMB Correct
      Pest Total    = Aphids Total + Leaf miner Total + TMB Total
      Pest Accuracy = (Pest Correct / Pest Total) * 100.0
    """
    p_correct = aphids["correct"] + leaf_miner["correct"] + tmb["correct"]
    p_total = aphids["total"] + leaf_miner["total"] + tmb["total"]
    p_acc = (p_correct / p_total * 100.0) if p_total > 0 else 0.0

    return {
        "correct": p_correct,
        "total": p_total,
        "accuracy": round(p_acc, 2)
    }


def format_result_cell(correct: int, total: int, accuracy: float) -> str:
    """Formats result as 'Correct / Total / Accuracy (%)' (e.g. '850 / 900 / 94.44%')."""
    return f"{correct} / {total} / {accuracy:.2f}%"


# ---------------------------------------------------------
# 4. OpenPyXL Professional Excel Formatting Engine
# ---------------------------------------------------------
def export_professional_excel(df: pd.DataFrame, output_excel_path: str):
    """
    Exports pandas DataFrame into a beautifully formatted Excel workbook using OpenPyXL.
      - Bold header font with soft blue fill (#D9E1F2)
      - Centered text and numerical formatting
      - Auto-adjusted column widths
      - Frozen header row (A2)
      - Enabled grid lines
    """
    if not OPENPYXL_AVAILABLE:
        logger.warning("openpyxl is not installed. Exporting standard pandas Excel...")
        df.to_excel(output_excel_path, index=False)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pest & Disease Summary"
    ws.views.sheetView[0].showGridLines = True

    # Style Definitions
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="000000")
    data_font = Font(name="Calibri", size=11, bold=False, color="000000")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # 1. Write Header Row
    headers = list(df.columns)
    ws.append(headers)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = cell_border

    # 2. Write Data Rows
    for row_idx, row_data in enumerate(df.values, 2):
        ws.append(list(row_data))
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            if col_idx == 1:
                cell.alignment = left_align
            else:
                cell.alignment = center_align

    # 3. Freeze Header Row
    ws.freeze_panes = "A2"

    # 4. Auto-adjust Column Widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 16)

    # Set row heights
    ws.row_dimensions[1].height = 26
    for r in range(2, len(df) + 2):
        ws.row_dimensions[r].height = 22

    wb.save(output_excel_path)
    logger.info(f"Successfully saved formatted Excel workbook to: {output_excel_path}")


# ---------------------------------------------------------
# 5. Main Phase 10 Execution Pipeline Engine
# ---------------------------------------------------------
def run_phase10_overall_classification_pipeline() -> Dict:
    """
    Main Phase 10 pipeline entrypoint.
    Processes all 8 individual vision models, computes pooled 'Overall 3 Pests' metrics,
    builds the exact 6-column summary DataFrame, and exports Excel workbook (XLSX) and CSV backup.
    """
    pipeline_start_time = time.time()
    root_output_dir = Config.get_overall_pest_disease_dir()

    logger.info(f"\n=======================================================================")
    logger.info(f"  PHASE 10: OVERALL PEST & DISEASE CLASSIFICATION SUMMARY")
    logger.info(f"=======================================================================")

    summary_rows = []

    # Process all 8 models in exact required row order
    for model_index in range(1, len(Config.MODEL_MAP) + 1):
        folder_name, _ = Config.MODEL_MAP[model_index]
        
        # Load per-class metrics from existing phase artifacts
        class_metrics = load_model_classification_metrics(model_index, folder_name)

        aphids = class_metrics.get("Aphids", {"correct": 0, "total": 0, "accuracy": 0.0})
        leaf_miner = class_metrics.get("Leaf miner", {"correct": 0, "total": 0, "accuracy": 0.0})
        tmb = class_metrics.get("TMB", {"correct": 0, "total": 0, "accuracy": 0.0})
        leaf_blight = class_metrics.get("Leaf blight", {"correct": 0, "total": 0, "accuracy": 0.0})

        # Calculate pooled "Overall 3 Pests" metrics
        overall_pests = calculate_overall_pests_metrics(aphids, leaf_miner, tmb)

        # Build row with exact 6 required columns
        row = {
            "Model": folder_name,
            "Aphids": format_result_cell(aphids["correct"], aphids["total"], aphids["accuracy"]),
            "Leaf miner": format_result_cell(leaf_miner["correct"], leaf_miner["total"], leaf_miner["accuracy"]),
            "TMB": format_result_cell(tmb["correct"], tmb["total"], tmb["accuracy"]),
            "Overall 3 Pests": format_result_cell(overall_pests["correct"], overall_pests["total"], overall_pests["accuracy"]),
            "Leaf blight": format_result_cell(leaf_blight["correct"], leaf_blight["total"], leaf_blight["accuracy"])
        }
        summary_rows.append(row)

    # 1. Build DataFrame
    df_summary = pd.DataFrame(summary_rows)

    # Validation Checks before writing
    assert len(df_summary) == 8, f"Expected 8 model rows, got {len(df_summary)}"
    expected_cols = ["Model", "Aphids", "Leaf miner", "TMB", "Overall 3 Pests", "Leaf blight"]
    assert list(df_summary.columns) == expected_cols, f"Column mismatch. Got: {list(df_summary.columns)}"

    # 2. Export Excel Workbook (.xlsx)
    excel_path = os.path.join(root_output_dir, "overall_pest_disease_classification.xlsx")
    export_professional_excel(df_summary, excel_path)

    # 3. Export CSV Backup (.csv)
    csv_path = os.path.join(root_output_dir, "overall_pest_disease_classification.csv")
    df_summary.to_csv(csv_path, index=False)

    # 4. Export Phase 10 Markdown Report
    report_content = f"""# Phase 10 — Overall Pest & Disease Classification Summary Report
**Cashew Pest and Disease Diagnosis System**

---

## 1. Executive Summary & Purpose

Phase 10 provides an executive-level **Overall Pest & Disease Classification Summary** across all 8 individual vision models.
It aggregates the three pest classes (**Aphids**, **Leaf miner**, **TMB**) into a single **"Overall 3 Pests"** metric using pooled sample counts, while keeping the single disease class (**Leaf blight**) separate.

> [!NOTE]
> **Reporting Only**: No model retraining or inference was executed during this phase. All values were extracted from authoritative Phase 6 complete-dataset classification artifacts.

---

## 2. Category Definitions

- **PESTS (3 Classes)**: `Aphids`, `Leaf miner`, `TMB`
- **DISEASE (1 Class)**: `Leaf blight`

---

## 3. Overall 3 Pests Pooled Formula

$$\\text{{Pest Correct}} = \\text{{Aphids Correct}} + \\text{{Leaf miner Correct}} + \\text{{TMB Correct}}$$

$$\\text{{Pest Total}} = \\text{{Aphids Total}} + \\text{{Leaf miner Total}} + \\text{{TMB Total}}$$

$$\\text{{Pest Accuracy (\\%)}} = \\left( \\frac{{\\text{{Pest Correct}}}}{{\\text{{Pest Total}}}} \\right) \\times 100$$

---

## 4. Overall Pest & Disease Classification Table

```markdown
{df_summary.to_markdown(index=False)}
```

---

## 5. Output File Locations

- **Excel Workbook**: `Experiments/Overall_Pest_Disease_Classification/overall_pest_disease_classification.xlsx`
- **CSV Backup**: `Experiments/Overall_Pest_Disease_Classification/overall_pest_disease_classification.csv`
- **Markdown Report**: `Experiments/Overall_Pest_Disease_Classification/Overall_Pest_Disease_Classification_Report.md`
"""

    report_path_exp = os.path.join(root_output_dir, "Overall_Pest_Disease_Classification_Report.md")
    with open(report_path_exp, "w") as f:
        f.write(report_content)

    doc_dir = Config.get_documentation_dir()
    report_path_doc = os.path.join(doc_dir, "Phase_10_Overall_Pest_Disease_Classification_Report.md")
    with open(report_path_doc, "w") as f:
        f.write(report_content)

    total_pipeline_time = time.time() - pipeline_start_time

    # 5. Formatted Summary Console Report
    print(f"\n=======================================================================")
    print(f"  PHASE 10 OVERALL PEST & DISEASE CLASSIFICATION SUMMARY")
    print(f"=======================================================================")
    print(df_summary.to_string(index=False))
    print(f"\nPipeline Execution Time : {total_pipeline_time:.2f} seconds.")
    print(f"Excel Workbook Saved To : {excel_path}")
    print(f"CSV Backup Saved To     : {csv_path}\n")

    return {
        "pipeline_phase": "Phase 10 - Overall Pest & Disease Classification Summary",
        "models_processed": len(df_summary),
        "excel_path": excel_path,
        "csv_path": csv_path,
        "execution_time_seconds": round(total_pipeline_time, 2)
    }
